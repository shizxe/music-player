import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

from config import (
    APP_DIR,
    BOOTSTRAP_APP_FILE,
    BOOTSTRAP_PROVIDERS_FILE,
    MIGRATIONS_DIR,
    PROVIDERS_FILE,
    ROUTES_API_FILE,
    TEMPLATES_DIR,
)
from utils import (
    ensure_directory,
    load_file_text,
    normalize_model_name,
    normalize_table_name,
    parse_fields,
    pluralize,
    route_name_for_table,
    singular_table_name,
    snake_case,
    unique_migration_filename,
    validation_rule,
)


@dataclass
class ModuleDefinition:
    model: str
    table: str
    module: str
    fields: list[dict[str, Any]]
    route: str


def error(message: str) -> None:
    print(f"Error: {message}")
    sys.exit(1)


def load_csv(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        error(f"Unable to decode CSV file: {path}")

    rows = []
    reader = csv.reader(text.splitlines())
    for row in reader:
        values = [str(value).strip() if value is not None else "" for value in row]
        if any(values):
            rows.append(values)
    return rows


def load_xlsx(path: Path) -> list[tuple[str, list[list[str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    workbook_rows: list[tuple[str, list[list[str]]]] = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() if value is not None else "" for value in row]
            if any(values):
                rows.append(values)
        if rows:
            workbook_rows.append((sheet.title, rows))
    return workbook_rows


def normalize_sql_type(raw_type: str) -> tuple[str, str]:
    value = raw_type.strip().lower()
    value = re.sub(r"\s+", " ", value)

    if value.startswith("varchar") or value.startswith("string"):
        return "string", "string"
    if value.startswith("char"):
        return "string", "string"
    if value.startswith("decimal") or value.startswith("numeric"):
        return "decimal", "decimal"
    if value.startswith("tinyint") or value.startswith("bool"):
        return "boolean", "boolean"
    if value.startswith("int") or value.startswith("integer"):
        return "integer", "integer"
    if value.startswith("bigint"):
        return "bigint", "bigInteger"
    if value.startswith("datetime") or value.startswith("timestamp"):
        return "datetime", "dateTime"
    if value.startswith("date"):
        return "date", "date"
    if value.startswith("time"):
        return "time", "time"
    if value.startswith("text"):
        return "text", "text"
    if value.startswith("json"):
        return "json", "json"
    if value.startswith("uuid"):
        return "uuid", "uuid"
    return "string", "string"


def load_input_sources(path: Path) -> list[tuple[str, list[list[str]]]]:
    if path.is_dir():
        sources: list[tuple[str, list[list[str]]]] = []
        files = sorted(
            file for file in path.rglob("*")
            if file.is_file() and file.suffix.lower() in {".xlsx", ".csv"}
        )
        if not files:
            error(f"Input directory does not contain any .xlsx or .csv files: {path}")
        for file in files:
            sources.extend(load_input_sources(file))
        return sources

    if path.suffix.lower() == ".xlsx":
        return load_xlsx(path)

    if path.suffix.lower() == ".csv":
        return [(path.name, load_csv(path))]

    error("Input file must be .xlsx or .csv")
    return []


def normalize_headers(headers: list[str]) -> list[str]:
    return [header.strip().lower() if header else "" for header in headers]


def normalized_header_key(value: str | None) -> str:
    if not value:
        return ""
    value = value.strip().lower()
    value = value.replace(".", "")
    value = value.replace("-", "_")
    value = value.replace(" ", "_")
    value = re.sub(r"_+", "_", value)
    return value


def find_column_index(headers: list[str], name: str) -> int | None:
    name = normalized_header_key(name)
    for index, header in enumerate(headers):
        if normalized_header_key(header) == name:
            return index
    return None


def is_structured_definition(headers: list[str]) -> bool:
    normalized = [normalized_header_key(header) for header in headers]
    return (
        "logical_name" in normalized or "logicalname" in normalized
    ) and (
        "physical_name" in normalized or "physicalname" in normalized
    ) and (
        "type" in normalized or "datatype" in normalized
    )


def parse_structured_field_row(row: list[str], indexes: dict[str, int]) -> dict[str, Any] | None:
    physical_name = row[indexes["physical_name"]] if indexes.get("physical_name") is not None and indexes["physical_name"] < len(row) else ""
    if not physical_name:
        return None

    name = snake_case(physical_name)
    if name == "id":
        return None

    raw_type = row[indexes["type"]] if indexes.get("type") is not None and indexes["type"] < len(row) else "string"
    raw_type = raw_type.strip()
    if not raw_type:
        raw_type = "string"

    length_value = row[indexes["length"]] if indexes.get("length") is not None and indexes["length"] < len(row) else ""
    precision_value = row[indexes["precision"]] if indexes.get("precision") is not None and indexes["precision"] < len(row) else ""
    raw_type_key, field_type = normalize_sql_type(raw_type)
    if length_value and precision_value:
        raw_type = f"{raw_type_key}({length_value},{precision_value})"
    elif length_value:
        raw_type = f"{raw_type_key}({length_value})"
    elif raw_type_key == "bigint" and "unsigned" in raw_type.lower():
        raw_type = "unsignedBigInteger"
    elif raw_type_key == "integer" and "unsigned" in raw_type.lower():
        raw_type = "unsignedInteger"

    required_value = row[indexes["required"]] if indexes.get("required") is not None and indexes["required"] < len(row) else ""
    required = not (required_value.strip().lower() in {"n", "no", "false", "0", "nullable", "null"})

    logical_name = row[indexes["logical_name"]] if indexes.get("logical_name") is not None and indexes["logical_name"] < len(row) else ""
    comment_text = logical_name.strip() if logical_name else ""

    modifiers: list[str] = []
    if not required:
        modifiers.append("nullable")

    constraint_value = ""
    if indexes.get("constraints") is not None and indexes["constraints"] < len(row):
        constraint_value = str(row[indexes["constraints"]]).strip().lower()
        if constraint_value in {"uk", "unique", "unique key"}:
            modifiers.append("unique")
        elif constraint_value in {"pk", "primary"}:
            modifiers.append("primary")

    if indexes.get("primary_key") is not None and indexes["primary_key"] < len(row):
        primary_key = str(row[indexes["primary_key"]]).strip().lower()
        if primary_key in {"y", "yes", "true", "1", "pk", "p"}:
            modifiers.append("primary")

    reference_table = None
    on_delete = None
    if "fk" in constraint_value:
        if name.endswith("_id"):
            reference_table = pluralize(singular_table_name(name[:-3]))
    if indexes.get("delete_constraints") is not None and indexes["delete_constraints"] < len(row):
        on_delete = str(row[indexes["delete_constraints"]]).strip().lower() or None

    return {
        "name": name,
        "type": field_type,
        "raw_type": raw_type,
        "nullable": not required,
        "modifiers": modifiers,
        "comment": comment_text,
        "reference_table": reference_table,
        "on_delete": on_delete,
        "constraint": constraint_value,
    }


def parse_definitions(path: Path) -> list[ModuleDefinition]:
    sources = load_input_sources(path)
    if not sources:
        error("Input file does not contain any rows")

    definitions: list[ModuleDefinition] = []
    for source_name, sheet_rows in sources:
        if not sheet_rows:
            continue

        headers = normalize_headers(sheet_rows[0])
        body_rows = sheet_rows[1:]
        model_index = find_column_index(headers, "model")
        table_index = find_column_index(headers, "table")
        fields_index = find_column_index(headers, "fields")

        if model_index is not None and table_index is not None and fields_index is not None:
            for row in body_rows:
                model_value = row[model_index] if model_index < len(row) else ""
                if not model_value:
                    continue

                model_name = normalize_model_name(model_value)
                table_value = row[table_index] if table_index < len(row) else None
                fields_value = row[fields_index] if fields_index < len(row) else None
                if isinstance(table_value, str):
                    table_value = table_value.strip()
                    if table_value == "":
                        table_value = None
                if isinstance(fields_value, str):
                    fields_value = fields_value.strip()

                table_name = normalize_table_name(table_value, model_name)
                field_definitions = parse_fields(fields_value)
                definitions.append(
                    ModuleDefinition(
                        model=model_name,
                        table=table_name,
                        module=model_name,
                        fields=field_definitions,
                        route=route_name_for_table(table_name),
                    )
                )
            continue

        if is_structured_definition(headers):
            header_map = {
                "logical_name": find_column_index(headers, "logical_name") or find_column_index(headers, "logical name"),
                "physical_name": find_column_index(headers, "physical_name") or find_column_index(headers, "physical name"),
                "type": find_column_index(headers, "type") or find_column_index(headers, "datatype"),
                "length": find_column_index(headers, "length") or find_column_index(headers, "size"),
                "precision": find_column_index(headers, "precision") or find_column_index(headers, "scale"),
                "required": find_column_index(headers, "required") or find_column_index(headers, "nullable"),
                "primary_key": find_column_index(headers, "primary_key") or find_column_index(headers, "pk"),
                "constraints": find_column_index(headers, "constraints") or find_column_index(headers, "constraint"),
                "delete_constraints": find_column_index(headers, "delete_constraints") or find_column_index(headers, "delete constraint"),
            }
            table_name = normalize_table_name(source_name, source_name)
            model_name = normalize_model_name(source_name)
            field_definitions = []
            for row in body_rows:
                field = parse_structured_field_row(row, header_map)
                if field is None:
                    continue
                field_definitions.append(field)

            if field_definitions:
                definitions.append(
                    ModuleDefinition(
                        model=model_name,
                        table=table_name,
                        module=model_name,
                        fields=field_definitions,
                        route=route_name_for_table(table_name),
                    )
                )
            continue

    if not definitions:
        error(f"No module definitions found in {path}")

    return definitions


def render_template(name: str, context: dict[str, Any]) -> str:
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATES_DIR)),
        autoescape=False,
        trim_blocks=True,
        lstrip_blocks=True,
    )
    template = env.get_template(name)
    return template.render(**context)


def write_file(path: Path, content: str, force: bool = False) -> bool:
    if path.exists() and not force:
        return False
    path.write_text(content, encoding="utf-8")
    return True


def build_column_mappings(fields: list[dict[str, Any]], composite_unique_fields: list[str] | None = None) -> list[dict[str, Any]]:
    columns = []
    for field in fields:
        raw_type = field["raw_type"].lower()
        name = field["name"]
        method = "string"
        args: list[str] = []

        if field.get("reference_table"):
            method = "foreignId"
        elif raw_type.startswith("unsignedbiginteger"):
            method = "unsignedBigInteger"
        elif raw_type.startswith("unsignedinteger"):
            method = "unsignedInteger"
        elif raw_type.startswith("bigint"):
            method = "bigInteger"
        elif raw_type == "email":
            method = "string"
        elif raw_type == "password":
            method = "string"
        elif raw_type in {"bool", "boolean"}:
            method = "boolean"
        elif raw_type in {"int", "integer"}:
            method = "integer"
        elif raw_type == "datetime":
            method = "dateTime"
        elif raw_type == "json":
            method = "json"
        elif raw_type == "uuid":
            method = "uuid"
        elif raw_type == "text":
            method = "text"
        elif raw_type == "date":
            method = "date"
        elif raw_type == "time":
            method = "time"
        elif raw_type.startswith("string("):
            method = "string"
            length_match = re.search(r"string\((\d+)", raw_type)
            if length_match:
                args = [length_match.group(1)]
        elif raw_type.startswith("char("):
            method = "char"
            length_match = re.search(r"char\((\d+)", raw_type)
            if length_match:
                args = [length_match.group(1)]
        elif raw_type.startswith("varchar("):
            method = "string"
            length_match = re.search(r"varchar\((\d+)", raw_type)
            if length_match:
                args = [length_match.group(1)]
        elif raw_type.startswith("decimal("):
            method = "decimal"
            param_match = re.search(r"decimal\((\d+),(\d+)", raw_type)
            if param_match:
                args = [param_match.group(1), param_match.group(2)]
            else:
                length_match = re.search(r"decimal\((\d+)", raw_type)
                if length_match:
                    args = [length_match.group(1)]

        declaration = f"$table->{method}('{name}'"
        if method == "foreignId":
            declaration = f"$table->{method}('{name}')"
        else:
            if args:
                if method == "decimal":
                    declaration += f", {args[0]}, {args[1]}" if len(args) > 1 else f", {args[0]}"
                else:
                    declaration += f", {args[0]}"
            declaration += ")"

        unique = "unique" in field.get("modifiers", [])
        if composite_unique_fields and len(composite_unique_fields) > 1 and name in composite_unique_fields:
            unique = False

        columns.append({
            "name": name,
            "declaration": declaration,
            "nullable": field.get("nullable", False),
            "unique": unique,
        })
    return columns


def build_rules(fields: list[dict[str, Any]], update: bool = False) -> list[dict[str, str]]:
    rules = []
    for field in fields:
        rule = validation_rule(field, update=update)
        if not rule:
            continue
        rules.append({"field": field["name"], "rule": rule})
    return rules


def compute_stub_context(module: ModuleDefinition) -> dict[str, Any]:
    rows = [field["name"] for field in module.fields]
    composite_unique_fields = [field["name"] for field in module.fields if "unique" in field.get("modifiers", []) and field.get("name")]
    columns = build_column_mappings(module.fields, composite_unique_fields)
    foreign_keys = []
    unique_indexes = []
    for field in module.fields:
        if field.get("reference_table"):
            foreign_keys.append({
                "column": field["name"],
                "references": "id",
                "table": field["reference_table"],
                "on_delete": field.get("on_delete") or "cascade",
            })

    if len(composite_unique_fields) >= 2:
        unique_indexes = [{"columns": composite_unique_fields, "name": f"{module.table}_uk"}]
    elif len(composite_unique_fields) == 1:
        unique_indexes = [{"columns": composite_unique_fields, "name": f"{module.table}_{composite_unique_fields[0]}_uk"}]

    return {
        "model": module.model,
        "module": module.module,
        "table": module.table,
        "fields": rows,
        "columns": columns,
        "store_rules": build_rules(module.fields, update=False),
        "update_rules": build_rules(module.fields, update=True),
        "route": module.route,
        "controller_namespace": f"App\\Http\\Controllers\\{module.module}",
        "foreign_keys": foreign_keys,
        "unique_indexes": unique_indexes,
    }


def generate_model(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Models" / f"{module.model}.php"
    ensure_directory(path.parent)
    content = render_template("model.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_controller(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Http" / "Controllers" / module.module / f"{module.model}Controller.php"
    ensure_directory(path.parent)
    content = render_template("controller.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_store_request(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Http" / "Requests" / module.module / f"Store{module.model}Request.php"
    ensure_directory(path.parent)
    content = render_template("store_request.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_update_request(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Http" / "Requests" / module.module / f"Update{module.model}Request.php"
    ensure_directory(path.parent)
    content = render_template("update_request.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_service_interface(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Services" / module.module / f"{module.model}ServiceInterface.php"
    ensure_directory(path.parent)
    content = render_template("service_interface.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_service(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Services" / module.module / f"{module.model}Service.php"
    ensure_directory(path.parent)
    content = render_template("service.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_repository_interface(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Repositories" / module.module / f"{module.model}RepositoryInterface.php"
    ensure_directory(path.parent)
    content = render_template("repository_interface.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_repository(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Repositories" / module.module / f"{module.model}Repository.php"
    ensure_directory(path.parent)
    content = render_template("repository.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def generate_resource(module: ModuleDefinition, force: bool) -> bool:
    path = APP_DIR / "Http" / "Resources" / module.module / f"{module.model}Resource.php"
    ensure_directory(path.parent)
    content = render_template("resource.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def find_existing_migration(table_name: str) -> Path | None:
    pattern = f"_create_{table_name}_table.php"
    for path in MIGRATIONS_DIR.glob(f"*{pattern}"):
        return path
    return None


def generate_migration(module: ModuleDefinition, force: bool) -> bool:
    ensure_directory(MIGRATIONS_DIR)
    existing = find_existing_migration(module.table)
    if existing and not force:
        return False
    if existing and force:
        path = existing
    else:
        filename = unique_migration_filename(module.table)
        path = MIGRATIONS_DIR / filename
    content = render_template("migration.stub", compute_stub_context(module))
    return write_file(path, content, force=force)


def ensure_repository_service_provider(bindings: list[dict[str, str]]) -> bool:
    ensure_directory(PROVIDERS_FILE.parent)
    if not PROVIDERS_FILE.exists():
        content = render_template("provider.stub", {"bindings": bindings})
        PROVIDERS_FILE.write_text(content, encoding="utf-8")
    else:
        append_provider_bindings(PROVIDERS_FILE, bindings)
    return True


def append_provider_bindings(path: Path, bindings: list[dict[str, str]]) -> None:
    content = load_file_text(path)
    modified = False
    for binding in bindings:
        marker = f"{binding['interface']}::class"
        if marker in content:
            continue
        injection = (
            "        $this->app->bind(\n"
            f"            {binding['interface']}::class,\n"
            f"            {binding['implementation']}::class\n"
            "        );\n\n"
        )
        start = content.find("public function register(): void")
        if start == -1:
            continue
        brace_index = content.find("{", start)
        if brace_index == -1:
            continue
        insert_index = find_closing_brace(content, brace_index)
        if insert_index == -1:
            continue
        content = content[:insert_index] + injection + content[insert_index:]
        modified = True
    if modified:
        path.write_text(content, encoding="utf-8")


def find_closing_brace(text: str, start_index: int) -> int:
    depth = 0
    for index, char in enumerate(text[start_index:], start=start_index):
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                return index
    return -1


def ensure_bootstrap_provider_registered() -> None:
    if not BOOTSTRAP_PROVIDERS_FILE.exists():
        return
    content = load_file_text(BOOTSTRAP_PROVIDERS_FILE)
    marker = "App\\Providers\\RepositoryServiceProvider::class"
    if marker in content:
        return
    insertion = "    App\\Providers\\RepositoryServiceProvider::class,\n"
    content = content.replace("AppServiceProvider::class,\n", f"AppServiceProvider::class,\n{insertion}")
    BOOTSTRAP_PROVIDERS_FILE.write_text(content, encoding="utf-8")


def ensure_api_routes_loaded() -> None:
    if not BOOTSTRAP_APP_FILE.exists():
        return
    content = load_file_text(BOOTSTRAP_APP_FILE)
    if "api: __DIR__.'/../routes/api.php'" in content:
        return
    old = "withRouting(\n        web: __DIR__.'/../routes/web.php',\n"
    if old in content:
        content = content.replace(
            old,
            "withRouting(\n        web: __DIR__.'/../routes/web.php',\n        api: __DIR__.'/../routes/api.php',\n",
        )
        BOOTSTRAP_APP_FILE.write_text(content, encoding="utf-8")


def ensure_api_route_file(module: ModuleDefinition) -> bool:
    if not ROUTES_API_FILE.exists():
        ensure_directory(ROUTES_API_FILE.parent)
        content = "<?php\n\nuse Illuminate\\Support\\Facades\\Route;\n\n"
        ROUTES_API_FILE.write_text(content, encoding="utf-8")

    content = load_file_text(ROUTES_API_FILE)
    controller_import = f"use App\\Http\\Controllers\\{module.module}\\{module.model}Controller;"
    route_definition = render_template(
        "route.stub",
        {
            "route": module.route,
            "controller": f"{module.model}Controller",
        },
    ).strip()
    modified = False
    if controller_import not in content:
        import_block = "use Illuminate\\Support\\Facades\\Route;\n"
        if import_block in content:
            content = content.replace(import_block, f"{import_block}{controller_import}\n")
        else:
            content = f"<?php\n\n{controller_import}\n\n" + content
        modified = True
    if route_definition not in content:
        if not content.endswith("\n"):
            content += "\n"
        content += f"{route_definition}\n"
        modified = True
    if modified:
        ROUTES_API_FILE.write_text(content, encoding="utf-8")
    return modified


def generate_module(module: ModuleDefinition, force: bool) -> dict[str, bool]:
    results = {}
    results["Model"] = generate_model(module, force)
    results["Controller"] = generate_controller(module, force)
    results["Store Request"] = generate_store_request(module, force)
    results["Update Request"] = generate_update_request(module, force)
    results["Service Interface"] = generate_service_interface(module, force)
    results["Service"] = generate_service(module, force)
    results["Repository Interface"] = generate_repository_interface(module, force)
    results["Repository"] = generate_repository(module, force)
    results["Resource"] = generate_resource(module, force)
    results["Migration"] = generate_migration(module, force)
    return results


def build_bindings(modules: list[ModuleDefinition]) -> list[dict[str, str]]:
    bindings: list[dict[str, str]] = []
    for module in modules:
        repo_interface = f"App\\Repositories\\{module.module}\\{module.model}RepositoryInterface"
        repo_impl = f"App\\Repositories\\{module.module}\\{module.model}Repository"
        service_interface = f"App\\Services\\{module.module}\\{module.model}ServiceInterface"
        service_impl = f"App\\Services\\{module.module}\\{module.model}Service"
        bindings.append({"interface": repo_interface, "implementation": repo_impl})
        bindings.append({"interface": service_interface, "implementation": service_impl})
    return bindings


def print_generation_result(module: ModuleDefinition, results: dict[str, bool]) -> None:
    print(f"Generating {module.model}...")
    for key, created in results.items():
        status = "Created" if created else "Skipped"
        print(f"{status} {key}")
    print("")


def main() -> None:
    parser = argparse.ArgumentParser(description="Laravel CRUD code generator")
    parser.add_argument("input", help="Path to modules .xlsx or .csv file")
    parser.add_argument("--force", "-f", action="store_true", help="Overwrite existing files")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        error(f"Input file not found: {input_path}")

    modules = parse_definitions(input_path)
    if not modules:
        error("No module definitions found in input file")

    ensure_bootstrap_provider_registered()
    ensure_api_routes_loaded()

    all_bindings = build_bindings(modules)
    ensure_repository_service_provider(all_bindings)
    ensure_bootstrap_provider_registered()

    for module in modules:
        results = generate_module(module, args.force)
        print_generation_result(module, results)
        ensure_api_route_file(module)

    print("Done.")


if __name__ == "__main__":
    main()
